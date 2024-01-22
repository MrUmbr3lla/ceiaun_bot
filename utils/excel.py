from openpyxl.reader.excel import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

import settings


def create_new_sheet(title: str) -> str:
    """
    Create new report from base excel template.

    Return report file path.
    """

    workbook = load_workbook(filename=settings.EXCEL_BASE_TEMPLATE)
    sheet = workbook.active
    sheet["B1"] = title

    file_path = settings.EXCEL_GENERATED_FILES_DIR / f"{title}.xlsx"
    workbook.save(filename=file_path)

    return file_path


def find_empty_row(sheet: Worksheet, start_row_number) -> int:
    """
    Find empty row number and return it.

    Note: Start with row number 3 by default.
    """

    row_number = start_row_number
    while True:
        if sheet[f"A{row_number}"].value is None:
            return row_number

        row_number += 1


def write_data_to_sheet(file_path: str, data: list, columns: list, start_row_number: int = 3) -> None:
    """
    Write one row data to first emprty row of sheet.

    Example:
         >>> write_data_to_sheet("example.xlsx", ["data-1", "data-2"], ["A", "B"])
         None
    """

    workbook = load_workbook(filename=file_path)
    sheet = workbook.active
    row_number = find_empty_row(sheet, start_row_number)

    for i, column in enumerate(columns):
        try:
            sheet[f"{column}{row_number}"] = data[i]
        except IndexError:
            pass

    workbook.save(filename=file_path)
